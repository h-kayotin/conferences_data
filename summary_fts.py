"""
summary_fts - 汇总代理商数据

Author: JiangHai江海
Date： 2023/4/25
"""

from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string as col_index
import datetime
import xlrd


def get_files_from_folder(input_path):
    folder_src = input_path  # type: Path
    files = folder_src.glob("*.*")
    return list(files)


def read_config():
    # config_dic = {
    #     "大连康乐美": {
    #         "sheet_name": "商品(进、销、存)滚动表",
    #         "data_start_col": 3,
    #         "static_start": "A",
    #         "static_end": "H",
    #         "data_field": [
    #             {
    #                 "router": "",
    #                 "start_row": "",
    #                 "end_row": ""
    #             }
    #         ]
    #     }
    # }

    config_dic = {}
    con_wb = openpyxl.load_workbook("sources/config.xlsx")  # type: Workbook

    # 读取各代理商表格参数
    con_sheet = con_wb.worksheets[0]  # type: Worksheet
    for row in range(2, con_sheet.max_row + 1):
        config_dic[con_sheet.cell(row, 1).value] = {
            "sheet_name": con_sheet.cell(row, 2).value,
            "data_start_row": int(con_sheet.cell(row, 3).value),
            "static_start": con_sheet.cell(row, 4).value,
            "static_end": con_sheet.cell(row, 5).value,
            "data_list": []
        }
        for col in range(6, con_sheet.max_column + 1):
            if con_sheet.cell(row, col).value is None:
                continue
            cell_values = con_sheet.cell(row, col).value.split(",")
            data = {
                "router": cell_values[0],
                "start_row": cell_values[1],
                "end_row": cell_values[2]
            }
            config_dic[con_sheet.cell(row, 1).value]["data_list"].append(data)

    # 读取表头
    title_sheet = con_wb.worksheets[1]  # type: Worksheet
    title_list = []
    for i in range(1, title_sheet.max_column + 1):
        title_list.append(title_sheet.cell(1, i).value)

    return config_dic, title_list


def read_xlsx(file_src, file_config, file_name):
    wb = openpyxl.load_workbook(file_src, data_only=True)  # type: Workbook
    sheet = wb[file_config["sheet_name"]]  # type: Worksheet

    start_row = file_config["data_start_row"]
    static_start = file_config["static_start"]
    static_end = file_config["static_end"]
    data_list = file_config["data_list"]

    processed_data_list = []
    for rout_index in range(len(data_list)):
        for row in range(start_row, sheet.max_row + 1):
            if type(sheet.cell(row, 1).value) is str:
                break
            # 如果销售额是0，跳过
            if sheet.cell(row, col_index(data_list[rout_index]["start_row"])).value == 0:
                continue

            processed_row = []
            for col in range(col_index(static_start), col_index(static_end) + 1):
                processed_row.append(sheet.cell(row, col).value)
            processed_row.append(sheet.cell(row, col_index(data_list[rout_index]["start_row"])).value)
            processed_row.append(sheet.cell(row, col_index(data_list[rout_index]["end_row"])).value)
            processed_row.append(data_list[rout_index]["router"])
            processed_row.append(datetime.datetime.today().strftime("%Y年%m月%d日"))
            processed_row.append(file_name)
            processed_row.append("一般贸易")
            processed_row.append("ST(销)")
            processed_data_list.append(processed_row)

    return processed_data_list


def read_xls(file_src, file_config, file_name):
    wb = xlrd.open_workbook(file_src)
    sheet = wb.sheet_by_name(file_config["sheet_name"])

    # xlrd的读取，是从0开始，所以都要减去1做下处理
    start_row = int(file_config["data_start_row"]) - 1
    static_start = col_index(file_config["static_start"]) - 1
    static_end = col_index(file_config["static_end"])
    data_list = file_config["data_list"]

    processed_data_list = []
    for rout_index in range(len(data_list)):
        for row in range(start_row, sheet.nrows):
            # print(type(sheet.cell(row, 0).value))
            if type(sheet.cell(row, 1).value) is str:
                break
            if sheet.cell(row, col_index(data_list[rout_index]["start_row"]) - 1).value == 0:
                continue

            processed_row = []
            for col in range(static_start, static_end):
                processed_row.append(sheet.cell(row, col).value)
            processed_row.append(sheet.cell(row, col_index(data_list[rout_index]["start_row"]) - 1).value)
            processed_row.append(sheet.cell(row, col_index(data_list[rout_index]["end_row"]) - 1).value)
            processed_row.append(data_list[rout_index]["router"])
            processed_row.append(datetime.datetime.today().strftime("%Y年%m月%d日"))
            processed_row.append(file_name)
            processed_row.append("一般贸易")
            processed_row.append("ST(销)")
            processed_data_list.append(processed_row)

    return processed_data_list


def write_sum(data_list, file_name, title_row):
    save_path = Path("./output/一般业务ST销售汇总.xlsx")
    if save_path.exists():
        sum_wb = openpyxl.load_workbook("./output/一般业务ST销售汇总.xlsx")
    else:
        sum_wb = openpyxl.Workbook()  # type: Workbook
    sum_sheet = sum_wb.worksheets[0]  # type: Worksheet

    sum_col = 1
    sum_row = sum_sheet.max_row + 1
    print(f"正在导入【{file_name}】的数据--->")

    if sum_row > 2:
        pass
    else:
        for i in range(1, len(title_row)+1):
            sum_sheet.cell(1, i, title_row[i-1])

    for row in range(1, len(data_list) + 1):
        for col in range(1, len(data_list[row-1]) + 1):
            data_row = data_list[row - 1]
            sum_sheet.cell(sum_row, sum_col, data_row[col-1])
            sum_col += 1
        sum_row += 1
        sum_col = 1

    print(f"共写入{len(data_list)}行数据--->\n")
    sum_wb.save("./output/一般业务ST销售汇总.xlsx")
    total_row = sum_sheet.max_row
    return total_row


def main():
    config_file, title_row = read_config()  # 读取配置文件
    while True:
        source_src = Path(input("请输入文件夹路径："))
        #  C:\Users\JiangHai江海\Desktop\工作\04.数据导出、合并\BI\source_BI
        files_list = get_files_from_folder(source_src)
        if source_src.exists():
            break
        else:
            print("路径有误，请重新输入--->\n")
    len_this, len_total = 0, 0
    for file in files_list:
        file_name = Path(file).stem
        file_type = str(file).split(".")[-1]
        if file_type == "xlsx":
            sum_list = read_xlsx(file, config_file[file_name], file_name)
        else:
            sum_list = read_xls(file, config_file[file_name], file_name)
        len_total = write_sum(sum_list, file_name, title_row)
        len_this += len(sum_list)
    return len_this, len_total


if __name__ == '__main__':
    len_t, len_sum = main()
    print(f"合并完毕，本次共合并{len_t}条数据，\n结果已保存到/output/一般业务ST销售汇总.xlsx")
    print(f"目前共有{len_sum}条数据------->\n")
    input("请按Enter键退出，或直接关闭程序")


