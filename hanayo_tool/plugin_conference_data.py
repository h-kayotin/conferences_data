"""
conference_263 - 读取各个会议平台的账单，写入到汇总表中
在之前公司使用的一个工具，现在估计没啥用了。放着参考吧、
这个只针对前公司的每个月合并几个数据表格文件
1. 账单和汇总表请放在同文件夹下的bills文件夹下面
2. 文件名请一定要和保持如下格式：

    263-2023-01.xlsx

    证通-2023-01.xlsx

    全时-2023-01.html

    loopup-2023-01.csv

    注意：以上名称请根据月份修改对应月份，比如2月是 263-2023-02.xlsx

    汇总表名称请保持不要修改：账单汇总

3. 使用方式

    运行plugin_conference_data.py程序即可

    然后按照提示输入文件名，注意此处只输入文件名，不要输入后缀
    比如导入证通的数据，请输入证通-2023-01
Author: ahjiang
Date 2023/3/20
"""

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv
from bs4 import BeautifulSoup
from openpyxl.styles import Font


def read_cells_value(start_row, sheet, col_user, col_money):
    """
    读取某个sheet中指定数据
    :param start_row: 数据开始的行号
    :param sheet: Worksheet对象
    :param col_user: username的列名
    :param col_money: 费用那一列
    :return: 返回读取到的数据列表
    """
    values_list = []
    for row_num in range(start_row, sheet.max_row+1):
        if type(sheet[f"{col_money}{row_num}"].value) == str:
            pass
        else:
            cell_value = {
                "username": sheet[f"{col_user}{row_num}"].value,
                "money": sheet[f"{col_money}{row_num}"].value
            }
            values_list.append(cell_value)
    return values_list


def clear_sheet_datas(sheet):
    """
    清除目标sheet的所有数据
    :param sheet: 目标sheet对象
    :return: 无返回值
    """
    for row in sheet.rows:
        for cell in row:
            cell.value = ""


def write_to_sheet(sheet, datas, start_row_num=1):
    """
    写入目标表格
    :param sheet: 目标表格对象
    :param datas: 要写入的数据
    :param start_row_num: 从第几行开始写入，默认从1
    :return: 返回该次写入数据的总金额
    """
    row_num = start_row_num
    sum_money = 0
    for data in datas:
        sheet[f"A{row_num}"] = data["username"]
        sheet[f"B{row_num}"] = data["money"]
        sum_money += data["money"]
        sheet[f"C{row_num}"] = f"=VLOOKUP(A{row_num},邮箱导出部门!A:C,3,FALSE)"
        sheet[f"D{row_num}"] = f"=VLOOKUP(C{row_num},部门对应成本中心!A:B,2,FALSE)"
        row_num += 1
    return sum_money


def write_into_sum(company, month, source_sheet, sheet_target_sum, total_amount):
    """
    写入账单金额到汇总sheet中
    :param company: 公司名
    :param month: 月份
    :param source_sheet:来源sheet对象
    :param sheet_target_sum: 目标sheet对象
    :param total_amount: 合计金额
    :return: 无返回值
    """
    if company == "263":
        if source_sheet == "账单金额-人工会":
            target_pos = "263-self"
        else:
            target_pos = "263-auto"
        sheet_target_sum[f"{target_months[int(month)]}{target_company[target_pos]}"] = total_amount
    else:
        sheet_target_sum[f"{target_months[int(month)]}{target_company[company]}"] = total_amount


def read_and_copy(wb, company, month, source_sheet, target_sheet_name, start_row, user_col, money_col):
    """
    读取和写入主程序
    :param wb:源文件工作簿对象
    :param company: 公司名称
    :param month: 账单月份
    :param source_sheet: 来源sheet
    :param target_sheet_name: 目标sheet
    :param start_row: 读取源数据开始的列
    :param user_col: 用户名所在列
    :param money_col: 账单金额所在列
    :return: 无返回值
    """
    sheet_source = wb[source_sheet]  # type:Worksheet
    # 读取账单数据
    values_source = read_cells_value(start_row, sheet_source, user_col, money_col)
    print(f"成功读取{company}会议{len(values_source)}条数据--->\n")
    # 读取汇总表格
    wb_target = openpyxl.load_workbook("../bills/账单汇总.xlsx")  # type:Workbook
    target_sheet = wb_target[target_sheet_name]  # type: Worksheet

    # 先清空旧数据
    clear_sheet_datas(target_sheet)
    print(f"已清除目标表格{target_sheet_name}sheet中的数据，即将开始写入--->\n")

    # 写入数据 目标sheet,同时计算各自总金额，留待填入汇总sheet
    sum_amount = write_to_sheet(target_sheet, values_source)

    # 读取费用总结sheet
    sheet_target_sum = wb_target["费用总结"]  # type:Worksheet

    # 根据月份写入到汇总sheet中
    write_into_sum(company, month, source_sheet, sheet_target_sum, sum_amount)

    # 修改汇总表中写的月份
    month_format(sheet_target_sum, month)

    # 保存数据
    wb_target.save("../bills/账单汇总.xlsx")
    print(f"成功写入{len(values_source)}条数据--->\n")

    return values_source


def month_format(sheet, month):
    """
    修改汇总表的月份为当前月份
    :param sheet: 目标sheet
    :param month: 当前月份
    :return: 无返回值
    """
    font_sty = Font(color="FF0000", bold=True, name="Calibri")
    month = int(month)
    sheet[f"{target_months[month]}1"].font = font_sty
    sheet["B18"] = f"{month}月"


def bill_data_main():
    # 输入Excel文件的名称，省略.xlsx
    wb_name = input("请输入数据源文件名称：")  # 263-2023-01
    # 记录账单的公司、年份和月份
    bill_company = wb_name.split("-")[0]
    bill_year = wb_name.split("-")[1]
    bill_month = wb_name.split("-")[2]
    print(f"即将读取{bill_company}的{bill_year}年{bill_month}月的账单--->\n")
    # 根据公司名字，判断如何导入
    if bill_company == "263":
        wb = openpyxl.load_workbook(f"bills/{wb_name}.xlsx", data_only=True)  # type:Workbook
        values_self = read_and_copy(wb, bill_company, bill_month, "账单金额-人工会", "263人工", 22, "B", "J")
        values_auto = read_and_copy(wb, bill_company, bill_month, "账单金额-自助会", "263自助", 22, "A", "I")
        # 将人工的数据也复制到自助里
        copy_target = openpyxl.load_workbook("../bills/账单汇总.xlsx")  # type:Workbook
        write_to_sheet(copy_target["263自助"], values_self, len(values_auto) + 1)
        copy_target.save("../bills/账单汇总.xlsx")
        print(f"成功将人工会议账单数据{len(values_self)}条写入自助sheet中--->\n")

    elif bill_company == "证通":
        wb = openpyxl.load_workbook(f"bills/{wb_name}.xlsx", data_only=True)  # type:Workbook
        read_and_copy(wb, bill_company, bill_month, "账单金额", "证通", 21, "A", "I")

    elif bill_company == "loopup":
        total_loop = 0
        with open(f"bills/{wb_name}.csv", encoding="utf-8") as file:
            loop_reader = csv.reader(file, delimiter=",", quotechar="'")
            total_row = 0
            for row_num, row in enumerate(loop_reader):
                total_row += 1
                if row_num == 0:
                    pass
                else:
                    total_loop += float(row[len(row) - 1])
        total_loop = int(total_loop * 100) / 100
        print(f"成功读取loopup{total_row}条数据，账单总额是{total_loop}美元，即将开始写入--->\n")
        wb_target_loop = openpyxl.load_workbook("../bills/账单汇总.xlsx")  # type:Workbook
        target_sheet_sum = wb_target_loop["费用总结"]  # type: Worksheet
        write_into_sum(bill_company, bill_month, "loopup", target_sheet_sum, total_loop)
        target_sheet_sum["B34"] = total_loop
        wb_target_loop.save("../bills/账单汇总.xlsx")
        print("写入成功--->")

    elif bill_company == "全时":
        # 全时需要借助BS库，读取HTML文件
        with open(f"bills/{wb_name}.html", "rb") as file:
            html = file.read()
            #  读取时需要指定解析器
            bs_html = BeautifulSoup(html, "html.parser")  # type:BeautifulSoup
            # 根据class找到对应的div块
            target_div = bs_html.find("div", attrs={'class': 'ownersummary'})
            # 再根据tr找到表格所有的行
            rows = target_div.find_all("tr")
            values_quanshi = []  # 保存需要的数据
            for row in rows:
                if len(row.find_all("td")) < 5:
                    pass
                else:
                    value = {
                        "username": row.find_all("td")[0].text,
                        "money": float(row.find_all("td")[5].text)
                    }
                    values_quanshi.append(value)
            print(f"成功获取到{bill_company}的{len(values_quanshi)}条数据,开始准备写入--->\n")
            # 获取目标工作簿和sheet
            workbook_target = openpyxl.load_workbook("../bills/账单汇总.xlsx")  # type:Workbook
            target_quanshi = workbook_target["全时"]  # type: Worksheet
            # 清理目标sheet源数据
            clear_sheet_datas(target_quanshi)
            # 开始写入到全时sheet中
            quanshi_sum = write_to_sheet(target_quanshi, values_quanshi)
            # 写入到汇总中
            target_sheet_sum = workbook_target["费用总结"]  # type: Worksheet
            write_into_sum(bill_company, bill_month, "全时", target_sheet_sum, quanshi_sum)
            workbook_target.save("../bills/账单汇总.xlsx")
            print("写入成功--->\n")


# 下面这俩是用来确认填入汇总sheet的位置的
target_months = "ABCDEFGHIJKLM"
target_company = {
    "全时": 2,
    "263-auto": 3,
    "263-self": 4,
    "证通": 5,
    "loopup": 9
}


if __name__ == '__main__':
    bill_data_main()
    is_over = input("是否需要继续导入，继续请输入y：")
    while is_over == "y":
        bill_data_main()
        is_over = input("是否需要继续导入，继续请输入y：")
    print("全部导入已结束--->")
