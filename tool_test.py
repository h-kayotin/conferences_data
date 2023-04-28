"""
test - 

Author: JiangHai江海
Date： 2023/4/28
"""

import hanayo_tool
from pathlib import Path
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

get_files_from_folder = hanayo_tool.hana_filename.get_files_from_folder
get_filenames = hanayo_tool.hana_filename.get_filenames

while True:
    source_src = Path(input("请输入文件夹路径："))
    files_list = get_files_from_folder(source_src)
    if source_src.exists():
        break
    else:
        print("路径有误，请重新输入--->\n")
# print(files_list)
# print(get_filenames(files_list))
uniq_list = []
for file in files_list:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb["销售"]  # type:Worksheet
    col_datas = set()
    for row in range(2, sheet.max_row + 1):
        col_datas.add(sheet.cell(row, 11).value)
    if len(col_datas) > 1:
        uniq_list.append(Path(file).stem)
print(uniq_list)


